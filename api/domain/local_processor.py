"""Local document processor — runs extraction without S3 or Postgres.

Processes files directly from the workspace filesystem and updates SQLite.
Respects PDF_BACKEND config and optional Mistral/LibreOffice backends.
"""

import asyncio
import json
import logging
import os
import shutil
import subprocess
import tempfile
import uuid
from pathlib import Path

import aiosqlite
from llmwiki_core.chunking import MIN_CHUNK_TOKENS

from config import settings
from domain.file_types import (
    EXTRACTION_TYPES, HTML_TYPES, IMAGE_TYPES, OFFICE_TYPES,
    PDF_TYPES, SIMPLE_TEXT_TYPES, SPREADSHEET_TYPES,
)
from domain.watcher import mark_written
from infra.db.sqlite import (
    SQLiteDocumentRepository,
    _write_lock,
    create_pool,
    serialized_write,
)
from infra.tasks import spawn_logged
from services.extracted_assets import build_pdf_image_assets

logger = logging.getLogger(__name__)

# Cap concurrent fire-and-forget extractions so a burst of dropped files can't
# spawn one LibreOffice/OCR job (and connection) per file at once.
# 每路提取 = 一个 LibreOffice 或 JVM 进程(CPU+内存双重开销),默认按核数
# 取半、上限 8;大批量导入可用环境变量 EXTRACT_CONCURRENCY 显式调高。
PROCESS_CONCURRENCY = settings.EXTRACT_CONCURRENCY or max(2, min(8, (os.cpu_count() or 4) // 2))
_process_semaphore = asyncio.Semaphore(PROCESS_CONCURRENCY)

# 失败满 N 次进入隔离:不再随启动自动重试(真坏的文件每次重启都会白跑
# 一遍 LibreOffice/JVM/OCR)。「重新识别」或文件内容变化时清零计数。
FAILED_RETRY_LIMIT = 3

# Backward-compatible name for tests and callers that observe the processor's
# heavy-write gate. It aliases the single API-local SQLite write lock.
_db_write_gate = _write_lock

# 排队中或处理中的文档 id:sweep 周期的自愈兜底(kick_pending_backlog)
# 据此跳过已有任务在飞的文档,避免每轮重复入队。误判无害 —— claim 是
# 原子的(status='pending' 才领走),重复入队只会空转一次。
_inflight: set[str] = set()


async def process_document(db: aiosqlite.Connection, doc_id: str, workspace: Path) -> None:
    """Atomically claim a pending document, then extract text, chunk, update index."""
    claim = await db.execute(
        "UPDATE documents SET status = 'processing', error_message = NULL, "
        "updated_at = datetime('now') WHERE id = ? AND status = 'pending'",
        (doc_id,),
    )
    await db.commit()
    if claim.rowcount == 0:
        return

    cursor = await db.execute(
        "SELECT filename, file_type, relative_path FROM documents WHERE id = ?",
        (doc_id,),
    )
    row = await cursor.fetchone()
    if not row:
        logger.warning("Document %s not found", doc_id[:8])
        return

    cols = [d[0] for d in cursor.description]
    doc = dict(zip(cols, row))

    file_type = doc["file_type"] or ""
    file_path = workspace / doc["relative_path"]

    if not file_path.is_file():
        await db.execute(
            "UPDATE documents SET status = 'failed', error_message = 'File not found', "
            "extraction_attempts = extraction_attempts + 1, "
            "updated_at = datetime('now') WHERE id = ?",
            (doc_id,),
        )
        await db.commit()
        return

    try:
        if file_type in SIMPLE_TEXT_TYPES:
            content = file_path.read_text(encoding="utf-8", errors="replace")
            await replace_text_content(db, doc_id, content)
        elif file_type in (PDF_TYPES | OFFICE_TYPES) and _sniff_html(file_path):
            # 爬虫常把网页(错误页/正文)直接存成 .pdf/.doc 扩展名;
            # 按真实内容走 HTML 解析,而不是让两个提取引擎双双失败
            logger.info("HTML disguised as .%s, parsing as HTML: %s",
                        file_type, doc["filename"])
            await _process_html(db, doc_id, file_path)
        elif file_type in PDF_TYPES:
            await _process_pdf(db, doc_id, file_path, workspace)
        elif file_type in OFFICE_TYPES:
            await _process_office(db, doc_id, file_path, workspace)
        elif file_type in SPREADSHEET_TYPES:
            await _process_spreadsheet(db, doc_id, file_path)
        elif file_type in IMAGE_TYPES:
            await _process_image(db, doc_id)
        elif file_type in HTML_TYPES:
            await _process_html(db, doc_id, file_path)
        else:
            await db.execute(
                "UPDATE documents SET status = 'ready', extraction_attempts = 0, updated_at = datetime('now') WHERE id = ?",
                (doc_id,),
            )
            await db.commit()

        logger.info("Processed %s: %s", doc["filename"], file_type)

    except Exception as e:
        error_msg = str(e)[:500]
        # 标失败本身也可能碰锁;异常若逃逸,文档会卡死在 processing 且
        # 不再被任何路径重试。重试几轮,仍不行则留给启动回收复位。
        for attempt in range(5):
            try:
                await db.execute(
                    "UPDATE documents SET status = 'failed', error_message = ?, "
                    "extraction_attempts = extraction_attempts + 1, "
                    "updated_at = datetime('now') WHERE id = ?",
                    (error_msg, doc_id),
                )
                await db.commit()
                break
            except Exception:
                if attempt == 4:
                    logger.error("Could not mark %s failed (db busy); startup reconcile will reset it",
                                 doc_id[:8])
                else:
                    await asyncio.sleep(2)
        logger.error("Failed to process %s: %s", doc["filename"], e)


async def process_document_isolated(workspace: Path, doc_id: str) -> None:
    """Process a document on its own connection so fire-and-forget tasks can't
    flush another writer's open transaction on a shared connection."""
    _inflight.add(doc_id)
    try:
        async with _process_semaphore:
            db = await create_pool(str(workspace / ".llmwiki" / "index.db"), init_schema=False)
            try:
                await process_document(db, doc_id, workspace)
            finally:
                await db.close()
    finally:
        _inflight.discard(doc_id)


async def chunk_text_document(db: aiosqlite.Connection, doc_id: str, content: str | None) -> None:
    """Chunk an already-extracted text document so it becomes full-text searchable."""
    await replace_text_content(db, doc_id, content)


async def _replace_text_content_on_connection(
    db: aiosqlite.Connection,
    doc_id: str,
    content: str | None,
    chunks: list,
    *,
    parser: str = "text",
    store_chunks=None,
) -> int:
    """Replace local text-derived rows using the caller's active transaction."""
    cursor = await db.execute("SELECT version FROM documents WHERE id = ?", (doc_id,))
    row = await cursor.fetchone()
    if row is None:
        raise LookupError(f"document {doc_id} not found for text replacement")
    next_version = int(row[0] or 0) + 1
    writer = store_chunks or _store_chunks
    await writer(db, doc_id, chunks, next_version)
    await db.execute(
        "UPDATE documents SET status = 'ready', extraction_attempts = 0, content = ?, "
        "page_count = 1, parser = ?, version = ?, error_message = NULL, "
        "updated_at = datetime('now') WHERE id = ?",
        (content, parser, next_version, doc_id),
    )
    return next_version


async def replace_text_content(
    db: aiosqlite.Connection,
    doc_id: str,
    content: str | None,
    *,
    parser: str = "text",
) -> int:
    """Atomically rebuild a local text document from its current source content."""
    from services.chunker import chunk_text

    chunks = chunk_text(content or "")
    async with serialized_write(db):
        version = await _replace_text_content_on_connection(
            db,
            doc_id,
            content,
            chunks,
            parser=parser,
        )
        await db.commit()
    return version


async def reconcile_workspace(db: aiosqlite.Connection, workspace: Path) -> None:
    """启动对账:把停机断点处的所有文档重新排队提取。

    覆盖四类断点(每类扫描独立容错,库忙不拖垮整轮):
    1. 卡在 processing —— 停机瞬间正在提取的;
    2. 从未提取过 —— `llmwiki init` 只入索引不提取的;
    3. 排队中丢失的 pending —— 停机时在信号量后排队的任务随进程消失,
       这类文档可能带有旧提取结果(如「重新识别」重置的源文档,parser
       已设、chunks 还在),对"从未提取"的查询不可见,必须整体接住;
    4. 曾失败 —— 每次启动重试一次,镜像修复后无需手工干预。

    统一去重后经并行通道重跑(受 _process_semaphore 限流,与运行期
    吞吐一致)—— 逐个串行会让重启后的大积压恢复慢一个数量级。
    """
    to_process: dict[str, None] = {}   # 有序去重

    try:
        inconsistent_ids = await _inconsistent_ready_document_ids(db)
        if inconsistent_ids:
            await db.executemany(
                "UPDATE documents SET status = 'pending', updated_at = datetime('now') "
                "WHERE id = ? AND status = 'ready'",
                [(doc_id,) for doc_id in inconsistent_ids],
            )
            await db.commit()
            logger.warning(
                "Reconcile: queued %d ready docs with stale derived versions",
                len(inconsistent_ids),
            )
        to_process.update(dict.fromkeys(inconsistent_ids))
    except Exception:
        logger.exception("Reconcile: derived-version scan failed")

    try:
        # 卡在 processing = 上次处理中进程中断(停机,或正是该文档拖垮了
        # 进程)。复位时计一次尝试:靠 except 路径计数对崩溃型文件无效
        # (进程死了异常没机会抛)。反复中断达上限的转入失败隔离,防止
        # OOM 型文件每次重启都再拖垮一次进程;正常成功会清零计数,停机
        # 误伤的无辜文档下次跑通即自愈。
        cursor = await db.execute(
            "SELECT id FROM documents WHERE status = 'processing' "
            "AND extraction_attempts + 1 >= ?", (FAILED_RETRY_LIMIT,))
        crashed_ids = [r[0] for r in await cursor.fetchall()]
        if crashed_ids:
            await db.executemany(
                "UPDATE documents SET status = 'failed', "
                "error_message = '提取反复中断(疑似压垮进程),已停止自动重试', "
                "extraction_attempts = extraction_attempts + 1, "
                "updated_at = datetime('now') WHERE id = ?",
                [(i,) for i in crashed_ids])
            logger.warning("Reconcile: quarantined %d docs that repeatedly died mid-extraction",
                           len(crashed_ids))
        cursor = await db.execute("SELECT id FROM documents WHERE status = 'processing'")
        stuck_ids = [r[0] for r in await cursor.fetchall()]
        if stuck_ids:
            await db.execute(
                "UPDATE documents SET status = 'pending', "
                "extraction_attempts = extraction_attempts + 1 WHERE status = 'processing'")
            logger.info("Reconcile: reset %d docs stuck in 'processing'", len(stuck_ids))
        if crashed_ids or stuck_ids:
            await db.commit()
        to_process.update(dict.fromkeys(stuck_ids))
    except Exception:
        logger.exception("Reconcile: stuck-processing scan failed")

    extract_ids: list[str] = []
    try:
        extract_ids = await _unchunked_extractable_ids(db)
        if extract_ids:
            await db.executemany(
                "UPDATE documents SET status = 'pending', updated_at = datetime('now') "
                "WHERE id = ?",
                [(i,) for i in extract_ids],
            )
            await db.commit()
        to_process.update(dict.fromkeys(extract_ids))
    except Exception:
        logger.exception("Reconcile: unchunked scan failed")

    try:
        cursor = await db.execute(
            "SELECT id FROM documents WHERE status = 'pending' AND source_kind != 'asset'")
        to_process.update(dict.fromkeys(r[0] for r in await cursor.fetchall()))
    except Exception:
        logger.exception("Reconcile: pending scan failed")

    try:
        placeholders = ",".join("?" for _ in EXTRACTION_TYPES)
        cursor = await db.execute(
            f"SELECT id FROM documents WHERE status = 'failed' AND source_kind != 'asset' "
            f"AND extraction_attempts < ? AND file_type IN ({placeholders})",
            (FAILED_RETRY_LIMIT, *EXTRACTION_TYPES),
        )
        failed_ids = [r[0] for r in await cursor.fetchall()]
        if failed_ids:
            await db.executemany(
                "UPDATE documents SET status = 'pending', error_message = NULL, "
                "updated_at = datetime('now') WHERE id = ? AND status = 'failed'",
                [(i,) for i in failed_ids],
            )
            await db.commit()
            logger.info("Reconcile: retrying %d previously failed docs", len(failed_ids))
        to_process.update(dict.fromkeys(failed_ids))
    except Exception:
        logger.exception("Reconcile: failed-docs scan failed")

    # 预登记 _inflight:排到后面的文档在真正开跑前也算"有主",sweep 的
    # 自愈兜底不会重复入队。固定数量搬运工消费队列,避免为几万积压一次
    # 性创建等量 task 对象。
    ids = list(to_process)
    _inflight.update(ids)
    next_idx = 0

    async def _drain() -> None:
        nonlocal next_idx
        while next_idx < len(ids):
            doc_id = ids[next_idx]
            next_idx += 1
            try:
                await process_document_isolated(workspace, doc_id)
            except Exception:
                logger.exception("Reconcile: processing %s failed", doc_id[:8])
            finally:
                _inflight.discard(doc_id)

    workers = [spawn_logged(_drain(), f"reconcile-drain-{n}")
               for n in range(min(PROCESS_CONCURRENCY, len(ids)))]

    if ids:
        logger.info(
            "Reconcile: %d docs queued for extraction (concurrency %d)",
            len(ids), PROCESS_CONCURRENCY,
        )
    if workers:
        await asyncio.gather(*workers, return_exceptions=True)


async def kick_pending_backlog(db: aiosqlite.Connection, workspace: Path) -> int:
    """自愈兜底(sweep 每轮调用):pending 且无在飞任务的文档重新入队。

    正常路径(watcher/上传/重新识别)spawn 的任务若因异常或重启丢失,
    文档会永远滞留 pending;这里把它们捞回并行通道。每轮限量防止极端
    情况下重复入队堆积 —— claim 原子性保证重复也只是空转。
    """
    cursor = await db.execute(
        "SELECT id FROM documents WHERE status = 'pending' AND source_kind != 'asset'")
    ids = [r[0] for r in await cursor.fetchall() if r[0] not in _inflight]
    for doc_id in ids[:200]:
        spawn_logged(process_document_isolated(workspace, doc_id), f"kick:{doc_id[:8]}")
    return len(ids[:200])


async def _store_chunks(
    db: aiosqlite.Connection,
    doc_id: str,
    chunks: list,
    document_version: int,
) -> None:
    """Store chunks into SQLite, replacing any existing ones.

    批量 executemany:OCR 长文档动辄数百 chunk,逐条 execute 的跨线程
    往返会显著拉长写闸门内的临界区。"""
    await db.execute("DELETE FROM document_chunks WHERE document_id = ?", (doc_id,))
    if chunks:
        await db.executemany(
            "INSERT INTO document_chunks (id, document_id, chunk_index, content, source_content, page, "
            "start_char, token_count, header_breadcrumb, document_version) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [(str(uuid.uuid4()), doc_id, c.index, c.content, c.content, c.page,
              c.start_char, c.token_count, c.header_breadcrumb, document_version)
             for c in chunks],
        )


# ── PDF extraction ────────────────────────────────────────────────────────

async def _save_local_images(
    db: aiosqlite.Connection, doc_id: str, workspace: Path,
    pages_with_images: list[tuple[int, str, list[dict]]],
) -> dict[int, dict]:
    """Save extracted images as hidden sibling assets and return page metadata.

    文件落盘(IO 段)在写闸门外完成;库操作合并为闸门内单事务:删旧
    资产 + 批量取号插入 + 父文档 metadata 更新,一次 commit。多图报告
    曾经每图一个事务(取号/插入/提交/整行回读),数百图即数百次独立
    提交,且走 Repository 锁而非写闸门,与页面/chunk 写入交叉争锁。
    """
    repo = SQLiteDocumentRepository(db)
    doc = await repo.get(doc_id)
    if not doc:
        return {}

    assets, page_elements = build_pdf_image_assets(
        doc_id,
        doc["filename"],
        doc["path"],
        pages_with_images,
    )
    if not assets:
        return {}

    relpaths = [(a.path.rstrip("/") + "/" + a.filename).lstrip("/") for a in assets]
    for asset, relative_asset in zip(assets, relpaths):
        local_asset = workspace / relative_asset
        local_asset.parent.mkdir(parents=True, exist_ok=True)
        mark_written(str(local_asset))
        local_asset.write_bytes(asset.data)

    asset_metadata = [a.metadata() for a in assets]   # 含 sha256,闸门外算完
    async with serialized_write(db):
        await db.execute(
            "DELETE FROM documents WHERE source_kind = 'asset' AND metadata LIKE ?",
            (f'%"parent_document_id": "{doc_id}"%',),
        )
        # 落盘→入闸门的窗口可远超 mark_written 的 2s 冷却:watcher/sweep 可能
        # 已把这些图片抢注成普通文档行,批量 INSERT 会撞 UNIQUE(relative_path)
        # 使提取确定性失败 —— 按 relative_path 清场兜底(同一事务内)
        await db.executemany(
            "DELETE FROM documents WHERE relative_path = ?",
            [(rp,) for rp in relpaths],
        )
        # DELETE 已持写锁,MAX 取号到批量插入之间不会有其他连接插队
        cursor = await db.execute("SELECT COALESCE(MAX(document_number), 0) FROM documents")
        base_number = (await cursor.fetchone())[0]
        await db.executemany(
            "INSERT INTO documents (id, user_id, filename, title, path, relative_path, "
            "source_kind, file_type, file_size, status, content, tags, version, "
            "document_number, metadata) "
            "VALUES (?, ?, ?, ?, ?, ?, 'asset', ?, ?, 'ready', NULL, '[]', 0, ?, ?)",
            [(a.document_id, doc["user_id"], a.filename, a.filename, a.path, rp,
              a.file_type, len(a.data), base_number + i + 1, json.dumps(m))
             for i, (a, rp, m) in enumerate(zip(assets, relpaths, asset_metadata))],
        )
        # 父文档 metadata 并入同一事务(等价 set_metadata_field,不单独提交)
        cursor = await db.execute("SELECT metadata FROM documents WHERE id = ?", (doc_id,))
        row = await cursor.fetchone()
        try:
            parent_meta = json.loads(row[0]) if row and row[0] else {}
        except (json.JSONDecodeError, TypeError):
            parent_meta = {}
        parent_meta["assets"] = asset_metadata
        await db.execute(
            "UPDATE documents SET metadata = ?, updated_at = datetime('now') WHERE id = ?",
            (json.dumps(parent_meta), doc_id),
        )
        await db.commit()
    return page_elements


async def _store_page_contents(
    db: aiosqlite.Connection, doc_id: str,
    page_contents: list[tuple[int, str]], parser: str,
    page_elements: dict[int, dict] | None = None,
) -> None:
    """Store extracted pages, chunks, and update document status."""
    num_pages = len(page_contents)

    full_content = "\n\n---\n\n".join(md for _, md in page_contents)
    from services.chunker import chunk_pages
    chunks = chunk_pages(page_contents)   # CPU 段在闸门外

    async with serialized_write(db):
        cursor = await db.execute("SELECT version FROM documents WHERE id = ?", (doc_id,))
        row = await cursor.fetchone()
        if row is None:
            raise LookupError(f"document {doc_id} not found for page replacement")
        next_version = int(row[0] or 0) + 1
        await db.execute("DELETE FROM document_pages WHERE document_id = ?", (doc_id,))
        if page_contents:
            await db.executemany(
                "INSERT INTO document_pages "
                "(id, document_id, page, content, elements, document_version) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                [(str(uuid.uuid4()), doc_id, page_num, content,
                  json.dumps((page_elements or {}).get(page_num))
                  if (page_elements or {}).get(page_num) else None,
                  next_version)
                 for page_num, content in page_contents],
            )
        await _store_chunks(db, doc_id, chunks, next_version)
        await db.execute(
            "UPDATE documents SET status = 'ready', extraction_attempts = 0, content = ?, page_count = ?, "
            "parser = ?, version = ?, error_message = NULL, updated_at = datetime('now') WHERE id = ?",
            (full_content, num_pages, parser, next_version, doc_id),
        )
        await db.commit()


_HTML_SIGNATURES = (b"<!doc", b"<html", b"<head", b"<body", b"<?xml")


def _sniff_html(file_path: Path) -> bool:
    """扩展名与内容不符的常见场景:网页被存成 .pdf/.doc 扩展名。"""
    try:
        with open(file_path, "rb") as fh:
            head = fh.read(512).lstrip()[:64].lower()
    except OSError:
        return False
    return head.startswith(_HTML_SIGNATURES) or b"<html" in head


async def _process_pdf(db: aiosqlite.Connection, doc_id: str, file_path: Path, workspace: Path) -> None:
    """Extract PDF text. Uses opendataloader by default, Mistral if configured."""
    if settings.PDF_BACKEND == "mistral" and settings.MISTRAL_API_KEY:
        await _process_pdf_mistral(db, doc_id, file_path, workspace)
    else:
        from services.pdf_extract import extract_pdf
        with tempfile.TemporaryDirectory() as tmpdir:
            # 短名软链再提取:opendataloader 以输入文件名派生输出名,
            # 统一短名规避超长/特殊字符文件名的边界问题
            src = Path(tmpdir) / "source.pdf"
            os.symlink(file_path.resolve(), src)
            pages_with_images = await asyncio.to_thread(extract_pdf, str(src))
        page_elements = await _save_local_images(db, doc_id, workspace, pages_with_images)
        page_contents = [(num, md) for num, md, _ in pages_with_images]
        await _store_page_contents(db, doc_id, page_contents, "opendataloader", page_elements)


# ── Office processing ─────────────────────────────────────────────────────

def _extract_docx_paragraphs(file_path: Path) -> list[str]:
    """docx zip 直读兜底:LibreOffice 对损坏/非标准 docx 可能直接崩溃
    (SIGABRT,重试无用),而 docx 本质是 zip 包 —— 解包 word/document.xml
    即可拿到全部段落文本。丢版面,但比整份失败强(与 PDF 的 pypdf 兜底
    同一哲学)。表格单元格内的段落同样被 w:p 遍历覆盖。"""
    import zipfile
    from xml.etree import ElementTree

    W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    with zipfile.ZipFile(file_path) as z:
        xml = z.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    paras: list[str] = []
    for p in root.iter(f"{W}p"):
        text = "".join(t.text or "" for t in p.iter(f"{W}t")).strip()
        if text:
            paras.append(text)
    if not paras:
        raise RuntimeError("docx 兜底未提取到文本")
    return paras


async def _process_office(db: aiosqlite.Connection, doc_id: str, file_path: Path, workspace: Path) -> None:
    """Convert Office docs to PDF via local LibreOffice, then extract text."""
    lo = shutil.which("libreoffice") or shutil.which("soffice")
    if not lo:
        await db.execute(
            "UPDATE documents SET status = 'failed', "
            "error_message = 'LibreOffice not installed. Install it to process Office files.', "
            "updated_at = datetime('now') WHERE id = ?",
            (doc_id,),
        )
        await db.commit()
        return

    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            # 短名软链再转换:让输出名可预期,规避超长/特殊字符文件名的边界问题
            ext = file_path.suffix.lstrip(".") or "doc"
            src = Path(tmpdir) / f"source.{ext}"
            os.symlink(file_path.resolve(), src)
            result = await asyncio.to_thread(
                subprocess.run,
                [lo, "--headless", "--norestore",
                 # 每次调用独立用户配置目录:并发转换共用 ~/.config/libreoffice
                 # 会互踩配置锁,后来者直接失败(批量导入时成片报错的根因)
                 f"-env:UserInstallation=file://{tmpdir}/lo-profile",
                 "--convert-to", "pdf", "--outdir", tmpdir, str(src)],
                capture_output=True, timeout=120,
            )
            if result.returncode != 0:
                raise RuntimeError(f"LibreOffice conversion failed: {result.stderr.decode()[:300]}")

            pdf_files = list(Path(tmpdir).glob("*.pdf"))
            if not pdf_files:
                # 载入失败(如加密/受保护文档)时 soffice 退出码为 0、原因只打在
                # stdout —— 带出来让失败原因可诊断
                detail = (result.stdout.decode(errors="replace")[:200]
                          or result.stderr.decode(errors="replace")[:200]).strip()
                raise RuntimeError(
                    f"LibreOffice produced no PDF output{(': ' + detail) if detail else ''}")

            converted_pdf = pdf_files[0]

            # Store converted PDF in cache for the viewer
            cache_dir = workspace / ".llmwiki" / "cache" / "local" / doc_id
            cache_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(converted_pdf, cache_dir / "converted.pdf")

            from services.pdf_extract import extract_pdf
            pages_with_images = await asyncio.to_thread(extract_pdf, str(converted_pdf))
        except (RuntimeError, subprocess.TimeoutExpired) as e:
            if file_path.suffix.lower() != ".docx":
                raise    # .doc 等二进制格式没有廉价兜底,照常失败
            try:
                paras = await asyncio.to_thread(_extract_docx_paragraphs, file_path)
            except Exception as fb:
                raise RuntimeError(f"{e}(docx 兜底亦失败: {fb})") from e
            logger.warning("LibreOffice failed for %s, used docx-xml fallback: %s",
                           file_path.name, str(e)[:200])
            await _store_page_contents(db, doc_id, [(1, "\n\n".join(paras))], "docx-xml")
            return

    page_elements = await _save_local_images(db, doc_id, workspace, pages_with_images)
    page_contents = [(num, md) for num, md, _ in pages_with_images]
    await _store_page_contents(db, doc_id, page_contents, "libreoffice+opendataloader", page_elements)


# ── Mistral OCR ───────────────────────────────────────────────────────────

async def _process_pdf_mistral(db: aiosqlite.Connection, doc_id: str, file_path: Path, workspace: Path) -> None:
    """Extract PDF via Mistral OCR API (better tables/layout, requires API key)."""
    import httpx
    import base64

    pdf_bytes = file_path.read_bytes()
    pdf_b64 = base64.b64encode(pdf_bytes).decode()

    async with httpx.AsyncClient(timeout=120) as client:
        resp = await client.post(
            "https://api.mistral.ai/v1/ocr",
            headers={"Authorization": f"Bearer {settings.MISTRAL_API_KEY}"},
            json={
                "model": "mistral-ocr-latest",
                "document": {"type": "document_url", "document_url": f"data:application/pdf;base64,{pdf_b64}"},
            },
        )
        resp.raise_for_status()
        result = resp.json()

    pages = result.get("pages", [])
    page_contents = [(i + 1, p.get("markdown", "")) for i, p in enumerate(pages)]
    await _store_page_contents(db, doc_id, page_contents, "mistral")


# ── Spreadsheet processing ────────────────────────────────────────────────

def _read_sheet_rows(file_path: Path) -> list[tuple[str, list[str]]]:
    """按表读出 (表名, 行文本列表)。老式 .xls 走 xlrd,其余走 openpyxl。"""
    if file_path.suffix.lower() == ".xls":
        import xlrd
        book = xlrd.open_workbook(str(file_path))
        return [
            (sheet.name,
             [" | ".join("" if c.value is None else str(c.value) for c in sheet.row(r))
              for r in range(sheet.nrows)])
            for sheet in book.sheets()
        ]
    from openpyxl import load_workbook
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        return [
            (name,
             [" | ".join(str(c) if c is not None else "" for c in row)
              for row in wb[name].iter_rows(values_only=True)])
            for name in wb.sheetnames
        ]
    finally:
        wb.close()


async def _process_spreadsheet(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Extract spreadsheet data (openpyxl / xlrd). Stores pages AND chunks for search."""
    sheets = await asyncio.to_thread(_read_sheet_rows, file_path)

    async with serialized_write(db):
        await _store_spreadsheet_rows(db, doc_id, sheets, file_path)


async def _store_spreadsheet_rows(db: aiosqlite.Connection, doc_id: str,
                                  sheets: list[tuple[str, list[str]]], file_path: Path) -> None:
    cursor = await db.execute("SELECT version FROM documents WHERE id = ?", (doc_id,))
    row = await cursor.fetchone()
    if row is None:
        raise LookupError(f"document {doc_id} not found for spreadsheet replacement")
    next_version = int(row[0] or 0) + 1
    await db.execute("DELETE FROM document_pages WHERE document_id = ?", (doc_id,))

    all_content = []
    page_contents = []
    page_rows = []
    for i, (sheet_name, rows) in enumerate(sheets, 1):
        content = "\n".join(rows)
        page_rows.append((str(uuid.uuid4()), doc_id, i, content,
                          json.dumps({"sheet_name": sheet_name}), next_version))
        all_content.append(f"## {sheet_name}\n\n{content}")
        page_contents.append((i, content))
    if page_rows:
        await db.executemany(
            "INSERT INTO document_pages "
            "(id, document_id, page, content, elements, document_version) "
            "VALUES (?, ?, ?, ?, ?, ?)", page_rows,
        )

    num_sheets = len(sheets)
    full_content = "\n\n".join(all_content)

    from services.chunker import chunk_pages
    chunks = chunk_pages(page_contents)
    await _store_chunks(db, doc_id, chunks, next_version)

    parser = "xlrd" if file_path.suffix.lower() == ".xls" else "openpyxl"
    await db.execute(
        "UPDATE documents SET status = 'ready', extraction_attempts = 0, content = ?, page_count = ?, "
        "parser = ?, version = ?, error_message = NULL, updated_at = datetime('now') WHERE id = ?",
        (full_content, num_sheets, parser, next_version, doc_id),
    )
    await db.commit()


# ── Image / HTML processing ──────────────────────────────────────────────

async def _process_image(db: aiosqlite.Connection, doc_id: str) -> None:
    """Images are stored as-is — just mark ready."""
    await db.execute(
        "UPDATE documents SET status = 'ready', extraction_attempts = 0, page_count = 1, "
        "parser = 'native', version = version + 1, error_message = NULL, "
        "updated_at = datetime('now') WHERE id = ?",
        (doc_id,),
    )
    await db.commit()


async def _process_html(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Extract HTML content via webmd parser."""
    raw_html = file_path.read_text(encoding="utf-8", errors="replace")

    try:
        from html_parser import Parser
        parser = Parser(raw_html, content_only=True)
        result = parser.parse()
        content = result.content
    except Exception:
        content = raw_html

    from services.chunker import chunk_text
    chunks = chunk_text(content)
    async with serialized_write(db):
        cursor = await db.execute("SELECT version FROM documents WHERE id = ?", (doc_id,))
        row = await cursor.fetchone()
        if row is None:
            raise LookupError(f"document {doc_id} not found for HTML replacement")
        next_version = int(row[0] or 0) + 1
        await _store_chunks(db, doc_id, chunks, next_version)
        await db.execute(
            "UPDATE documents SET status = 'ready', extraction_attempts = 0, content = ?, page_count = 1, "
            "parser = 'webmd', version = ?, error_message = NULL, updated_at = datetime('now') "
            "WHERE id = ?",
            (content, next_version, doc_id),
        )
        await db.commit()


# ── Reconciliation queries ────────────────────────────────────────────────

async def _inconsistent_ready_document_ids(db: aiosqlite.Connection) -> list[str]:
    """Return ready non-assets whose derived chunks do not match their version."""
    cursor = await db.execute(
        "SELECT d.id FROM documents d "
        "WHERE d.status = 'ready' AND d.source_kind != 'asset' AND ("
        "EXISTS (SELECT 1 FROM document_chunks c "
        "        WHERE c.document_id = d.id AND c.document_version != d.version) "
        "OR (length(trim(COALESCE(d.content, ''))) >= ? AND NOT EXISTS "
        "    (SELECT 1 FROM document_chunks c WHERE c.document_id = d.id)))",
        (MIN_CHUNK_TOKENS * 4,),
    )
    return [row[0] for row in await cursor.fetchall()]


async def _unchunked_extractable_ids(db: aiosqlite.Connection) -> list[str]:
    """IDs of never-processed extractable docs (PDF/Office/spreadsheet/HTML) with no chunks.

    Excludes 'processing' so reconcile never reclaims a doc an isolated task is mid-extracting.
    """
    placeholders = ",".join("?" for _ in EXTRACTION_TYPES)
    cursor = await db.execute(
        f"SELECT id FROM documents WHERE status NOT IN ('failed', 'processing') AND source_kind != 'asset' "
        f"AND parser IS NULL "
        f"AND file_type IN ({placeholders}) "
        f"AND id NOT IN (SELECT DISTINCT document_id FROM document_chunks)",
        tuple(EXTRACTION_TYPES),
    )
    return [r[0] for r in await cursor.fetchall()]


async def _unchunked_text_docs(db: aiosqlite.Connection) -> list[tuple[str, str]]:
    """(id, content) for never-chunked simple-text docs that have content."""
    placeholders = ",".join("?" for _ in SIMPLE_TEXT_TYPES)
    cursor = await db.execute(
        f"SELECT id, content FROM documents WHERE status NOT IN ('failed', 'processing') AND source_kind != 'asset' "
        f"AND parser IS NULL "
        f"AND file_type IN ({placeholders}) "
        f"AND content IS NOT NULL AND content != '' "
        f"AND id NOT IN (SELECT DISTINCT document_id FROM document_chunks)",
        tuple(SIMPLE_TEXT_TYPES),
    )
    return [(r[0], r[1]) for r in await cursor.fetchall()]
